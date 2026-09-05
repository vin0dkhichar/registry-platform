import csv
import importlib
import importlib.util
import io
import os
import sys
import tempfile
import unittest
import zipfile
from datetime import timedelta
from pathlib import Path
from types import ModuleType, SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook
from sqlalchemy import String
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column


def _stub_iam_data_policy_helper():
    try:
        from iam_core.helpers.data_policy_helper import DataPolicyHelper  # noqa: F401
        return
    except ModuleNotFoundError:
        pass

    import iam_core

    helpers = sys.modules.get("iam_core.helpers") or ModuleType("iam_core.helpers")
    helpers.__path__ = []  # type: ignore[attr-defined]
    sys.modules["iam_core.helpers"] = helpers
    iam_core.helpers = helpers
    helper_mod = ModuleType("iam_core.helpers.data_policy_helper")

    class DataPolicyHelper:
        @staticmethod
        def resolve_register_record_policy(*_args, **_kwargs):
            return None

    helper_mod.DataPolicyHelper = DataPolicyHelper
    sys.modules["iam_core.helpers.data_policy_helper"] = helper_mod
    helpers.data_policy_helper = helper_mod


_stub_iam_data_policy_helper()
sys.modules.setdefault(
    "openg2p_registry_extensions",
    importlib.import_module(
        os.environ.get(
            "REGISTRY_EXTENSION_MODULE",
            "openg2p_registry_farmer_extension",
        )
    ),
)


def _load_register_export_worker_module():
    module_name = "openg2p_registry_celery_worker.tasks.register_export_worker"
    existing = sys.modules.get(module_name)
    if existing is not None and hasattr(existing, "RegisterExportWriter"):
        return existing

    import openg2p_registry_celery_worker

    tasks_name = "openg2p_registry_celery_worker.tasks"
    if tasks_name not in sys.modules:
        tasks_pkg = ModuleType(tasks_name)
        tasks_pkg.__path__ = [
            str(
                Path(openg2p_registry_celery_worker.__file__).resolve().parent
                / "tasks"
            )
        ]
        sys.modules[tasks_name] = tasks_pkg
        openg2p_registry_celery_worker.tasks = tasks_pkg

    worker_path = (
        Path(openg2p_registry_celery_worker.__file__).resolve().parent
        / "tasks"
        / "register_export_worker.py"
    )
    spec = importlib.util.spec_from_file_location(module_name, worker_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


_export_worker = _load_register_export_worker_module()
RegisterExportWriter = _export_worker.RegisterExportWriter
_remove_seen_rows = _export_worker._remove_seen_rows
_upload_export_output = _export_worker._upload_export_output
register_export_worker = _export_worker.register_export_worker

from openg2p_registry_core.models import (
    DocumentBucket,
    ExportFormatEnum,
    ProcessStatusEnum,
)


class Base(DeclarativeBase):
    pass


class ExportRecord(Base):
    __tablename__ = "test_export_record"

    internal_record_id: Mapped[str] = mapped_column(
        String, primary_key=True
    )
    display_name: Mapped[str] = mapped_column(String)


class RegisterExportWriterTests(unittest.TestCase):
    def setUp(self):
        self.register = SimpleNamespace(
            register_id="people",
            register_mnemonic="People",
        )
        self.hierarchy = SimpleNamespace(
            ordered_registers=[self.register]
        )
        self.models = {"people": ExportRecord}
        self.record = ExportRecord(
            internal_record_id="one",
            display_name="=HYPERLINK(\"https://example.test\")",
        )

    def test_writes_xlsx_and_escapes_formula_values(self):
        writer = RegisterExportWriter(
            self.hierarchy, self.models, ExportFormatEnum.XLSX
        )
        try:
            writer.write_rows("people", [self.record])
            output_path, extension, _ = writer.finalize()
            workbook = load_workbook(output_path, read_only=True)
            rows = list(workbook["People"].iter_rows(values_only=True))
            self.assertEqual(extension, "xlsx")
            self.assertEqual(
                rows[0], ("internal_record_id", "display_name")
            )
            self.assertTrue(rows[1][1].startswith("'="))
        finally:
            writer.close()

    def test_writes_zip_csv_and_escapes_formula_values(self):
        writer = RegisterExportWriter(
            self.hierarchy, self.models, ExportFormatEnum.ZIP_CSV
        )
        try:
            writer.write_rows("people", [self.record])
            output_path, extension, _ = writer.finalize()
            with zipfile.ZipFile(output_path) as archive:
                csv_name = archive.namelist()[0]
                csv_text = archive.read(csv_name).decode("utf-8")
            rows = list(csv.reader(io.StringIO(csv_text)))
            self.assertEqual(extension, "zip")
            self.assertTrue(rows[1][1].startswith("'="))
        finally:
            writer.close()

    def test_unique_sheet_names_and_xlsx_row_limit(self):
        first = SimpleNamespace(register_id="one", register_mnemonic="People")
        second = SimpleNamespace(register_id="two", register_mnemonic="People")
        hierarchy = SimpleNamespace(ordered_registers=[first, second])
        models = {"one": ExportRecord, "two": ExportRecord}
        writer = RegisterExportWriter(
            hierarchy, models, ExportFormatEnum.XLSX
        )
        try:
            writer.write_rows("one", [self.record])
            writer._row_counts["two"] = 1_048_575
            with self.assertRaisesRegex(ValueError, "ZIP_CSV"):
                writer.write_rows("two", [self.record])
            output_path, _, _ = writer.finalize()
            workbook = load_workbook(output_path, read_only=True)
            self.assertEqual(workbook.sheetnames, ["People", "People_1"])
        finally:
            writer.close()


class RegisterExportWorkerLogicTests(unittest.TestCase):
    def test_remove_seen_rows_deduplicates_across_batches(self):
        first = SimpleNamespace(internal_record_id="a")
        duplicate = SimpleNamespace(internal_record_id="a")
        second = SimpleNamespace(internal_record_id="b")
        seen_ids = {"people": set()}

        first_batch = _remove_seen_rows(
            {"people": [first, second]}, seen_ids
        )
        second_batch = _remove_seen_rows(
            {"people": [duplicate, second]}, seen_ids
        )

        self.assertEqual(first_batch["people"], [first, second])
        self.assertEqual(second_batch["people"], [])

    def test_upload_export_output_uses_named_object_and_timedelta_expiry(self):
        handler = MagicMock()
        handler.upload.return_value = "register-exports/export-1.xlsx"
        handler.get_url.return_value = "https://files.test/export-1.xlsx"

        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            temp_file.write(b"xlsx-bytes")
            output_path = temp_file.name
        try:
            with patch(
                "openg2p_registry_celery_worker.tasks.register_export_worker.get_document_handler",
                return_value=handler,
            ), patch(
                "openg2p_registry_celery_worker.tasks.register_export_worker._config"
            ) as config:
                config.export_files_prefix = "register-exports/"
                config.export_presigned_url_expiry_hours = 48
                object_name, url, expires_at = _upload_export_output(
                    output_path,
                    "export-1",
                    "xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        finally:
            os.unlink(output_path)

        self.assertEqual(object_name, "register-exports/export-1.xlsx")
        self.assertEqual(url, "https://files.test/export-1.xlsx")
        self.assertIsNotNone(expires_at)
        upload_args, upload_kwargs = handler.upload.call_args
        self.assertEqual(upload_args[2], DocumentBucket.EXPORT_FILES)
        self.assertEqual(
            upload_kwargs["object_name"], "register-exports/export-1.xlsx"
        )
        url_args, url_kwargs = handler.get_url.call_args
        self.assertEqual(url_args[0], "register-exports/export-1.xlsx")
        self.assertEqual(url_args[1], DocumentBucket.EXPORT_FILES)
        self.assertEqual(url_kwargs["expires"], timedelta(seconds=48 * 3600))

    def _session_for_queue_item(self, queue_item):
        session = MagicMock()
        session.get.return_value = queue_item
        session_cm = MagicMock()
        session_cm.__enter__.return_value = session
        session_cm.__exit__.return_value = False
        session_factory = MagicMock(return_value=session_cm)
        return session, session_factory

    def test_worker_skips_completed_exports(self):
        queue_item = SimpleNamespace(
            export_status=ProcessStatusEnum.COMPLETED.value,
            export_no_of_attempts=1,
        )
        session, session_factory = self._session_for_queue_item(queue_item)
        with patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker.sessionmaker",
            return_value=session_factory,
        ), patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker._process_export"
        ) as process_export:
            register_export_worker("export-1")

        process_export.assert_not_called()
        session.commit.assert_not_called()

    def test_worker_returns_pending_exports_to_queue_until_max_attempts(self):
        queue_item = SimpleNamespace(
            export_id="export-1",
            export_status=ProcessStatusEnum.PROCESSING.value,
            export_no_of_attempts=0,
            export_latest_timestamp=None,
            export_latest_error_code=None,
            last_processed_offset=12,
        )
        session, session_factory = self._session_for_queue_item(queue_item)
        with patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker.sessionmaker",
            return_value=session_factory,
        ), patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker._process_export",
            side_effect=RuntimeError("generation failed"),
        ), patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker._config"
        ) as config:
            config.export_worker_max_attempts = 3
            register_export_worker("export-1")

        self.assertEqual(queue_item.export_no_of_attempts, 1)
        self.assertEqual(
            queue_item.export_status, ProcessStatusEnum.PENDING.value
        )
        self.assertEqual(queue_item.last_processed_offset, 0)
        self.assertIn("generation failed", queue_item.export_latest_error_code)

    def test_worker_marks_failed_after_max_attempts(self):
        queue_item = SimpleNamespace(
            export_id="export-1",
            export_status=ProcessStatusEnum.PROCESSING.value,
            export_no_of_attempts=2,
            export_latest_timestamp=None,
            export_latest_error_code=None,
            last_processed_offset=0,
        )
        session, session_factory = self._session_for_queue_item(queue_item)
        with patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker.sessionmaker",
            return_value=session_factory,
        ), patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker._process_export",
            side_effect=RuntimeError("generation failed"),
        ), patch(
            "openg2p_registry_celery_worker.tasks.register_export_worker._config"
        ) as config:
            config.export_worker_max_attempts = 3
            register_export_worker("export-1")

        self.assertEqual(queue_item.export_no_of_attempts, 3)
        self.assertEqual(
            queue_item.export_status, ProcessStatusEnum.FAILED.value
        )
        session.rollback.assert_called()

